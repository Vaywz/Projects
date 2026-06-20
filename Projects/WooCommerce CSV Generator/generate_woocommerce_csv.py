#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Pierre Cardin Bedding — WooCommerce CSV Generator v2
Parses PDF price list + image folders → produces import-ready CSV/XLSX
"""

import argparse
import csv
import re
import logging
from pathlib import Path
from collections import defaultdict

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─── CONFIG ──────────────────────────────────────────────────────────────────
DEFAULT_PDF_PATH = Path("PierreCardinShop2026.pdf")
DEFAULT_IMAGES_ROOT = Path("extracted_images") / "Pierre Cardin 2026" / "Pierre Cardin 2026"
DEFAULT_OUTPUT_DIR = Path("output")
DEFAULT_IMAGE_BASE_URL = "https://pierrecardinbedding.lv/wp-content/uploads/products"

IMAGE_BASE_URL = DEFAULT_IMAGE_BASE_URL

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Generate WooCommerce CSV/XLSX imports from a PDF price list and product image folders."
    )
    parser.add_argument(
        "--pdf",
        type=Path,
        default=DEFAULT_PDF_PATH,
        help=f"Path to the source PDF price list. Default: {DEFAULT_PDF_PATH}",
    )
    parser.add_argument(
        "--images-root",
        type=Path,
        default=DEFAULT_IMAGES_ROOT,
        help=f"Path to the root folder that contains product image folders. Default: {DEFAULT_IMAGES_ROOT}",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Directory for generated CSV/XLSX/report files. Default: {DEFAULT_OUTPUT_DIR}",
    )
    parser.add_argument(
        "--image-base-url",
        default=DEFAULT_IMAGE_BASE_URL,
        help=f"Public base URL used for WooCommerce image fields. Default: {DEFAULT_IMAGE_BASE_URL}",
    )
    return parser.parse_args()

# ─── COLLECTION MAP ──────────────────────────────────────────────────────────
COLLECTION_MAP = {
    "Platinum": "Platinum",
    "Platinim": "Platinum",
    "Gold": "Gold",
    "Silver": "Silver",
}

# ─── MATTRESS-ONLY MODELS (standalone mattresses, not bed+mattress sets) ─────
MATTRESS_ONLY_MODELS = {
    "ORGANIC LATEX ROLLPACK", "COMFORT LATEX ROLLPACK",
    "COLL MEMORY VISCO ROLLPACK", "DUAL LUXE", "BIO LATEX",
    "VISCO BALANCE", "COTTON WOOL", "POWER ATHLETIC", "WOOLLIA",
    "LA TEXIA", "COOL WARM", "HAPPY NEST", "MOON BABY",
}

MATTRESS_TYPE_MAP = {
    "ORGANIC LATEX ROLLPACK": "Latex",
    "COMFORT LATEX ROLLPACK": "Latex",
    "BIO LATEX": "Latex",
    "LA TEXIA": "Latex",
    "COLL MEMORY VISCO ROLLPACK": "Memory Foam",
    "VISCO BALANCE": "Memory Foam",
    "DUAL LUXE": "Spring",
    "COTTON WOOL": "Spring",
    "POWER ATHLETIC": "Spring",
    "WOOLLIA": "Spring",
    "COOL WARM": "Spring",
    "HAPPY NEST": "Children",
    "MOON BABY": "Children",
}

# ─── IMAGE FOLDER → PDF MODEL MAPPING ────────────────────────────────────────
# key = folder name in ZIP, value = list of PDF model names that use these images
FOLDER_TO_MODELS = {
    "Afrodit":        ["AFRODIT"],
    "Alicante 1":     ["ALICANTE", "ALICANTE PRIME"],
    "Alicante 2":     ["ALICANTE", "ALICANTE PRIME"],  # variant 2
    "Aroma Theraphy":  ["AROMA THERAPY"],
    "Bellagio":       ["BELLAGIO"],
    "Biolatex":       ["BIO LATEX"],
    "Bleo Marine":    [],
    "Brugge":         [],
    "Cabestan":       [],
    "Casablanca":     ["CASABLANCA"],
    "Centaure":       ["CENTAURE"],
    "Clermont":       [],
    "Cool Warm":      ["COOL WARM"],
    "Cordoba":        ["CORDOBA"],
    "Cotton Wool":    ["COTTON WOOL"],
    "Dual Luxe":      ["DUAL LUXE"],
    "Excalibur":      [],
    "Ibiza":          [],
    "La Fayete 1":    ["LA FAYETTE", "LA FAYETTE PRIME"],
    "La Fayete 2":    ["LA FAYETTE", "LA FAYETTE PRIME"],
    "Laspalmas 1":    ["LAS PALMAS", "LAS PALMAS PRIME"],
    "Laspalmas 2":    ["LAS PALMAS", "LAS PALMAS PRIME"],
    "Latexia":        ["LA TEXIA"],
    "Leon":           ["LEON"],
    "Limoges":        [],
    "Malaga":         [],
    "Marseille":      [],
    "Melodie":        ["MELODIE"],
    "Modena":         ["MODENA", "MODENA PRIME"],
    "Montreul":       [],
    "Napoli":         ["NAPOLI"],
    "Orleans 1":      [],
    "Orleans 2":      [],
    "Oviedo":         ["OVIEDO"],
    "Palazio":        ["PALAZIO"],
    "Parma":          ["PARMA"],
    "Strasbourg":     [],
    "Toledo":         ["TOLEDO"],
    "Verona":         ["VERONA"],
    "Visco Balance":  ["VISCO BALANCE"],
    "Voyage":         ["VOYAGE"],
    "Woollia":        ["WOOLLIA"],
}

# Reverse: PDF model → best image folder
def build_model_to_folder():
    m2f = {}
    for folder, models in FOLDER_TO_MODELS.items():
        for m in models:
            if m not in m2f:
                m2f[m] = folder
    return m2f

MODEL_TO_FOLDER = build_model_to_folder()


# ─── PDF PARSER ──────────────────────────────────────────────────────────────
def parse_pdf(pdf_path):
    """
    Parse the entire PDF and return a list of model dicts:
    {name, collection, sizes: [{size, bed_price, mattress_price, set_price}], accessories: {NIGHTSTAND: price, PUFF: price}}
    """
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                full_text += t + "\n"

    lines = full_text.split("\n")
    models = []
    cur = None

    def save_cur():
        if cur and (cur["sizes"] or cur["accessories"]):
            models.append(cur)

    for raw in lines:
        line = raw.strip()
        if not line or line.startswith("PRICE LIST"):
            continue

        # ── detect collection tag ────────────────────────────────────
        for tag, val in COLLECTION_MAP.items():
            if line == tag:
                if cur:
                    cur["collection"] = val
                line = ""
                break
            if line.endswith(" " + tag):
                if cur:
                    cur["collection"] = val
                line = line[: -len(tag)].strip()
                break

        if not line:
            continue

        # ── detect dual-header FIRST: "- NAME1 - - NAME2 -" (mattress pages)
        dual_hdr = re.match(r"^-\s*(.+?)\s*-\s+-\s*(.+?)\s*-\s*$", line)
        if dual_hdr:
            save_cur()
            # We'll handle dual columns specially below; for now start with left model
            cur = {"name": dual_hdr.group(1).strip(), "collection": None, "sizes": [], "accessories": {},
                   "_right": dual_hdr.group(2).strip()}
            continue

        # ── single model header: "- NAME -" ─────────────────────────
        hdr = re.match(r"^-\s*(.+?)\s*-\s*$", line)
        if hdr:
            save_cur()
            cur = {"name": hdr.group(1).strip(), "collection": None, "sizes": [], "accessories": {}}
            continue

        # bare upper-case name not caught above (e.g. "MELODIE")
        if (line.isupper() and len(line) > 3 and "EUR" not in line
                and line not in ("PUFF", "NIGTHSTAND", "NIGHTSTAND", "NIGTHSTAND")
                and "Headboard" not in line and "Size" not in line and "Mattress" not in line):
            save_cur()
            cur = {"name": line.strip(), "collection": None, "sizes": [], "accessories": {}}
            continue

        if cur is None:
            continue

        # ── accessories ──────────────────────────────────────────────
        pf = re.match(r"PUFF\s+EUR\s+([\d\s]+)", line)
        if pf:
            cur["accessories"]["PUFF"] = int(pf.group(1).replace(" ", ""))
            continue
        ns = re.match(r"NI(?:GH?T|GTH)STAND\s+EUR\s+([\d\s]+)", line)
        if ns:
            cur["accessories"]["NIGHTSTAND"] = int(ns.group(1).replace(" ", ""))
            continue

        # ── bed price row  SIZE  EUR bed  EUR mattress  EUR set ──────
        bp = re.match(r"(\d+X\d+)\s+EUR\s+([\d ]+?)EUR\s+([\d ]+?)EUR\s+([\d ]+)", line)
        if bp:
            cur["sizes"].append({
                "size": bp.group(1),
                "bed_price": int(bp.group(2).replace(" ", "")),
                "mattress_price": int(bp.group(3).replace(" ", "")),
                "set_price": int(bp.group(4).replace(" ", "")),
            })
            continue

        # ── dual-column mattress row  SIZE EUR p1  SIZE EUR p2 ───────
        dm = re.match(r"(\d+X\d+)\s+EUR\s+([\d ]+?)\s+(\d+X\d+)\s+EUR\s+([\d ]+)", line)
        if dm:
            cur["sizes"].append({
                "size": dm.group(1),
                "mattress_price": int(dm.group(2).replace(" ", "")),
            })
            # right model
            right_name = cur.get("_right")
            if right_name:
                # stash right-side price for later
                cur.setdefault("_right_sizes", []).append({
                    "size": dm.group(3),
                    "mattress_price": int(dm.group(4).replace(" ", "")),
                })
            continue

        # ── single mattress row  SIZE EUR price ──────────────────────
        sm = re.match(r"(\d+X\d+)\s+EUR\s+([\d ]+)$", line)
        if sm:
            cur["sizes"].append({
                "size": sm.group(1),
                "mattress_price": int(sm.group(2).replace(" ", "")),
            })
            continue

    save_cur()

    # ── split dual-header models into two separate entries ────────────
    extra = []
    for m in models:
        right = m.pop("_right", None)
        right_sizes = m.pop("_right_sizes", [])
        if right and right_sizes:
            extra.append({
                "name": right,
                "collection": m["collection"],
                "sizes": right_sizes,
                "accessories": {},
            })
    models.extend(extra)

    return models


# ─── IMAGE SCANNER ───────────────────────────────────────────────────────────
def scan_images(root):
    """Return image_map[folder] = [Path, ...], ai_bgs[norm_name] = Path, yatak[norm] = [Path]"""
    image_map = {}
    ai_bgs = {}
    yatak = defaultdict(list)

    if not root.exists():
        return image_map, ai_bgs, yatak

    for folder in sorted(root.iterdir()):
        if not folder.is_dir():
            continue
        name = folder.name

        if name == "Ai Background":
            for f in folder.iterdir():
                if f.is_file() and f.suffix.lower() == ".webp":
                    key = re.sub(r"\s*(Ai|Al)$", "", f.stem, flags=re.I).strip().lower()
                    ai_bgs[key] = f
            continue

        if name == "yatak isimleri":
            for f in folder.iterdir():
                if f.is_file() and f.suffix.lower() == ".webp":
                    key = re.sub(r"\d+$", "", f.stem).strip().lower()
                    yatak[key].append(f)
            continue

        imgs = sorted(
            [f for f in folder.iterdir() if f.is_file() and f.suffix.lower() in (".webp", ".jpg", ".png")],
            key=lambda p: p.name
        )
        if imgs:
            image_map[name] = imgs

    return image_map, ai_bgs, yatak
   
def pick_images(model_name, image_map, ai_bgs):
    """Return (main_img, gallery_list, ai_bg) for a PDF model name."""
    folder = MODEL_TO_FOLDER.get(model_name)
    imgs = image_map.get(folder, []) if folder else []

    # Filter out AI images from the product folder
    product_imgs = [i for i in imgs if not re.search(r"\bAi\b", i.stem, re.I)]
    ai_in_folder = [i for i in imgs if re.search(r"\bAi\b", i.stem, re.I)]

    # AI background: prefer from Ai Background folder, fallback to one inside product folder
    norm = model_name.lower().replace(" ", "").replace("prime", "").strip()
    ai_bg = None
    for k, v in ai_bgs.items():
        if norm.replace(" ", "") in k.replace(" ", ""):
            ai_bg = v
            break
    if not ai_bg and ai_in_folder:
        ai_bg = ai_in_folder[0]

    main_img = product_imgs[0] if product_imgs else None
    gallery = product_imgs[1:6] if len(product_imgs) > 1 else []

    return main_img, gallery, ai_bg


# ─── HELPERS ─────────────────────────────────────────────────────────────────
def img_url(p):
    return f"{IMAGE_BASE_URL}/{p.parent.name}/{p.name}" if p else ""


def images_str(main, gallery, ai_bg=None):
    urls = []
    if main:
        urls.append(img_url(main))
    if ai_bg:
        urls.append(img_url(ai_bg))
    for g in gallery:
        u = img_url(g)
        if u not in urls:
            urls.append(u)
    return ", ".join(urls)


def sku(prefix, model, size="", variant=""):
    """Generate unique SKU like PC-BED-MODENA-160x200"""
    n = model.upper().replace(" ", "-")
    n = re.sub(r"[^A-Z0-9-]", "", n)
    parts = [f"PC-{prefix}", n]
    if variant:
        parts.append(variant)
    if size:
        parts.append(size.replace("X", "x"))
    return "-".join(parts)


def attr_cols(attrs):
    """attrs = [(name, value), ...] → dict of Attribute N ... columns"""
    out = {}
    for i, (name, val) in enumerate(attrs, 1):
        if val:
            out[f"Attribute {i} name"] = name
            out[f"Attribute {i} value(s)"] = str(val)
            out[f"Attribute {i} visible"] = 1
            out[f"Attribute {i} global"] = 1
    return out


# ─── BUILD PRODUCTS ─────────────────────────────────────────────────────────
def build_all(pdf_models, image_map, ai_bgs, yatak):
    products = []
    seen_skus = set()
    missing_img = []
    missing_price = []
    manual_rev = []
    img_map = []

    def add(p):
        s = p["SKU"]
        if s in seen_skus:
            # Append a suffix to avoid dupe
            for suffix in range(2, 100):
                candidate = f"{s}-V{suffix}"
                if candidate not in seen_skus:
                    p["SKU"] = candidate
                    s = candidate
                    break
        seen_skus.add(s)
        products.append(p)

    for md in pdf_models:
        name = md["name"]
        coll = md["collection"] or ""
        sizes = md["sizes"]
        acc = md["accessories"]
        is_mattress_only = name in MATTRESS_ONLY_MODELS or (sizes and "bed_price" not in sizes[0] and "set_price" not in sizes[0])
        base = name.replace(" PRIME", "").strip()
        is_prime = "PRIME" in name

        main_img, gallery, ai_bg = pick_images(name, image_map, ai_bgs)

        # Also try yatak images for mattresses
        yatak_imgs = []
        for yk, yv in yatak.items():
            if base.lower().replace(" ", "") in yk.replace(" ", ""):
                yatak_imgs.extend(yv)

        has_images = bool(main_img)

        if has_images:
            img_map.append({"model": name, "folder": MODEL_TO_FOLDER.get(name, ""), "images": len(gallery) + 1, "ai": "Yes" if ai_bg else "No"})

        badge = "Premium" if is_prime else ""

        # ── MATTRESS-ONLY ────────────────────────────────────────────
        if is_mattress_only:
            mtype = MATTRESS_TYPE_MAP.get(name, "")
            m_main = main_img or (yatak_imgs[0] if yatak_imgs else None)
            m_gal = gallery if gallery else yatak_imgs[:3]
            is_child = name in ("HAPPY NEST", "MOON BABY")

            if not m_main:
                missing_img.append({"model": name, "type": "mattress", "reason": "No images found"})

            for row in sizes:
                sz = row["size"]
                price = row.get("mattress_price", "")
                attrs_list = [
                    ("Size", sz.replace("X", "x")),
                    ("Collection", coll),
                    ("Mattress Type", mtype),
                    ("Availability", "Available"),
                ]
                if is_child:
                    attrs_list.append(("Product Badge", "Children"))

                p = {
                    "Type": "simple",
                    "SKU": sku("MAT", name, sz),
                    "Name": f"Pierre Cardin {name} Mattress {sz.replace('X','x')}",
                    "Published": 1,
                    "Regular price": price,
                    "Categories": "Mattresses",
                    "Brands": "Pierre Cardin",
                    "Short description": f"Pierre Cardin {name} mattress, {sz.replace('X','x')} cm. {mtype}. {coll} collection." if coll else f"Pierre Cardin {name} mattress, {sz.replace('X','x')} cm. {mtype}.",
                    "Description": f"Premium {name} mattress by Pierre Cardin. Size {sz.replace('X','x')} cm. {mtype} construction for optimal comfort.",
                    "Images": images_str(m_main, m_gal) if m_main else "",
                    "Meta: family_code": base.lower().replace(" ", "_"),
                    "Meta: collection": coll,
                    "Meta: configuration": "mattress",
                    "Meta: country_scope": "LV,LT,EE",
                }
                p.update(attr_cols(attrs_list))
                add(p)
            continue

        # ── BED MODEL (has bed_price / set_price) ────────────────────
        if not has_images:
            missing_img.append({"model": name, "type": "bed/set", "reason": "No image folder found"})

        for row in sizes:
            sz = row["size"]
            bed_price = row.get("bed_price", "")
            set_price = row.get("set_price", "")

            # — BED (without mattress) —
            if bed_price:
                attrs_list = [
                    ("Size", sz.replace("X", "x")),
                    ("Collection", coll),
                    ("Configuration", "Without Mattress"),
                    ("Mattress Included", "No"),
                    ("Availability", "Available"),
                ]
                if badge:
                    attrs_list.append(("Product Badge", badge))

                p = {
                    "Type": "simple",
                    "SKU": sku("BED", name, sz),
                    "Name": f"Pierre Cardin {name} Bed {sz.replace('X','x')}",
                    "Published": 1,
                    "Regular price": bed_price,
                    "Categories": "Beds",
                    "Brands": "Pierre Cardin",
                    "Short description": f"Pierre Cardin {name} bed frame {sz.replace('X','x')} cm. {coll} collection." if coll else f"Pierre Cardin {name} bed frame {sz.replace('X','x')} cm.",
                    "Description": f"Elegant {name} bed by Pierre Cardin. Size {sz.replace('X','x')} cm. Part of the {coll} collection." if coll else f"Elegant {name} bed by Pierre Cardin. Size {sz.replace('X','x')} cm.",
                    "Images": images_str(main_img, gallery, ai_bg),
                    "Meta: family_code": base.lower().replace(" ", "_"),
                    "Meta: collection": coll,
                    "Meta: configuration": "without_mattress",
                    "Meta: country_scope": "LV,LT,EE",
                }
                p.update(attr_cols(attrs_list))
                add(p)

            # — BEDROOM SET (with mattress) —
            if set_price:
                attrs_list = [
                    ("Size", sz.replace("X", "x")),
                    ("Collection", coll),
                    ("Configuration", "With Mattress"),
                    ("Mattress Included", "Yes"),
                    ("Availability", "Available"),
                ]
                if badge:
                    attrs_list.append(("Product Badge", badge))

                p = {
                    "Type": "simple",
                    "SKU": sku("SET", name, sz),
                    "Name": f"Pierre Cardin {name} Bedroom Set {sz.replace('X','x')}",
                    "Published": 1,
                    "Regular price": set_price,
                    "Categories": "Bedroom Sets",
                    "Brands": "Pierre Cardin",
                    "Short description": f"Pierre Cardin {name} bedroom set with mattress {sz.replace('X','x')} cm. {coll} collection." if coll else f"Pierre Cardin {name} bedroom set with mattress {sz.replace('X','x')} cm.",
                    "Description": f"Complete {name} bedroom set by Pierre Cardin. Bed frame + premium mattress, size {sz.replace('X','x')} cm.",
                    "Images": images_str(main_img, gallery, ai_bg),
                    "Meta: family_code": base.lower().replace(" ", "_"),
                    "Meta: collection": coll,
                    "Meta: configuration": "with_mattress",
                    "Meta: country_scope": "LV,LT,EE",
                }
                p.update(attr_cols(attrs_list))
                add(p)

        # ── ACCESSORIES ──────────────────────────────────────────────
        if "NIGHTSTAND" in acc:
            p = {
                "Type": "simple",
                "SKU": sku("NS", base),
                "Name": f"Pierre Cardin {base} Nightstand",
                "Published": 1,
                "Regular price": acc["NIGHTSTAND"],
                "Categories": "Accessories",
                "Brands": "Pierre Cardin",
                "Short description": f"Pierre Cardin {base} nightstand. {coll} collection." if coll else f"Pierre Cardin {base} nightstand.",
                "Description": f"Elegant nightstand from the {base} collection by Pierre Cardin.",
                "Images": "",
                "Meta: family_code": base.lower().replace(" ", "_"),
                "Meta: collection": coll,
                "Meta: configuration": "nightstand",
                "Meta: country_scope": "LV,LT,EE",
            }
            p.update(attr_cols([("Accessory type", "Nightstand"), ("Collection", coll), ("Availability", "Available")]))
            add(p)
            manual_rev.append({"SKU": p["SKU"], "model": base, "type": "Nightstand", "reason": "Needs dedicated photo"})

        if "PUFF" in acc:
            p = {
                "Type": "simple",
                "SKU": sku("PF", base),
                "Name": f"Pierre Cardin {base} Puff",
                "Published": 1,
                "Regular price": acc["PUFF"],
                "Categories": "Accessories",
                "Brands": "Pierre Cardin",
                "Short description": f"Pierre Cardin {base} puff ottoman. {coll} collection." if coll else f"Pierre Cardin {base} puff ottoman.",
                "Description": f"Stylish puff ottoman from the {base} collection by Pierre Cardin.",
                "Images": "",
                "Meta: family_code": base.lower().replace(" ", "_"),
                "Meta: collection": coll,
                "Meta: configuration": "puff",
                "Meta: country_scope": "LV,LT,EE",
            }
            p.update(attr_cols([("Accessory type", "Puff"), ("Collection", coll), ("Availability", "Available")]))
            add(p)
            manual_rev.append({"SKU": p["SKU"], "model": base, "type": "Puff", "reason": "Needs dedicated photo"})

    # ── MODELS WITH IMAGES BUT NO PDF PRICE (drafts) ─────────────────
    priced_bases = set()
    for md in pdf_models:
        priced_bases.add(md["name"])
        priced_bases.add(md["name"].replace(" PRIME", "").strip())

    for folder, models in FOLDER_TO_MODELS.items():
        if not models and folder in image_map:
            imgs = image_map[folder]
            product_imgs = [i for i in imgs if not re.search(r"\bAi\b", i.stem, re.I)]
            ai_bg = None
            for k, v in ai_bgs.items():
                if folder.lower().replace(" ", "") in k.replace(" ", ""):
                    ai_bg = v
                    break
            m = product_imgs[0] if product_imgs else None
            g = product_imgs[1:6] if len(product_imgs) > 1 else []

            p = {
                "Type": "simple",
                "SKU": sku("BED", folder),
                "Name": f"Pierre Cardin {folder} Bed",
                "Published": 0,
                "Regular price": "",
                "Categories": "Beds",
                "Brands": "Pierre Cardin",
                "Short description": f"Pierre Cardin {folder} bed. Contact us for pricing.",
                "Description": f"Pierre Cardin {folder} bed. Contact us for pricing and availability.",
                "Images": images_str(m, g, ai_bg),
                "Meta: family_code": folder.lower().replace(" ", "_"),
                "Meta: collection": "",
                "Meta: configuration": "",
                "Meta: country_scope": "LV,LT,EE",
            }
            add(p)
            missing_price.append({"model": folder, "type": "bed", "reason": "Has images, no PDF price"})

    return products, missing_img, missing_price, manual_rev, img_map


# ─── EXPORT ──────────────────────────────────────────────────────────────────
def all_columns(products):
    prio = ["Type", "SKU", "Name", "Published", "Regular price",
            "Categories", "Brands", "Short description", "Description", "Images"]
    extra = set()
    for p in products:
        extra.update(p.keys())
    attr_c = sorted(c for c in extra if c.startswith("Attribute"))
    meta_c = sorted(c for c in extra if c.startswith("Meta:"))
    rest = sorted(c for c in extra if c not in prio and c not in attr_c and c not in meta_c)
    return prio + rest + attr_c + meta_c


def export_csv(products, path, delim=","):
    cols = all_columns(products)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=cols, delimiter=delim, extrasaction="ignore")
        w.writeheader()
        for p in products:
            w.writerow(p)
    log.info(f"CSV → {path}  ({len(products)} rows)")


def export_xlsx(products, mi, mp, mr, im, path):
    wb = openpyxl.Workbook()
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    def sheet(ws, data, title):
        ws.title = title
        if not data:
            ws.append(["No data"])
            return
        keys = list(data[0].keys())
        ws.append(keys)
        for c in ws[1]:
            c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center")
        for r in data:
            ws.append([r.get(k, "") for k in keys])
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 2, 60)

    cols = all_columns(products)
    rows = [{c: p.get(c, "") for c in cols} for p in products]
    sheet(wb.active, rows, "final_import")
    sheet(wb.create_sheet(), im or [{"info": "—"}], "image_mapping")
    sheet(wb.create_sheet(), mr or [{"info": "—"}], "manual_review")
    sheet(wb.create_sheet(), mp or [{"info": "—"}], "missing_prices")
    sheet(wb.create_sheet(), mi or [{"info": "—"}], "missing_images")

    ws = wb.create_sheet("notes")
    for r in [
        ["Pierre Cardin Bedding — WooCommerce Import"],
        ["Generated", "2026-03-26"],
        [], ["Notes:"],
        [f"Image base URL (placeholder): {IMAGE_BASE_URL}"],
        ["Upload images to WP Media and replace URLs"],
        ["Published=0 → draft (no price)"],
        ["Check manual_review for items needing photos"],
        ["PRIME = same bed frame + premium mattress in set"],
    ]:
        ws.append(r)

    wb.save(path)
    log.info(f"XLSX → {path}")


def write_report_csv(data, path):
    if not data:
        return
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=data[0].keys())
        w.writeheader()
        w.writerows(data)


# ─── VALIDATION ──────────────────────────────────────────────────────────────
def validate(products):
    issues = []
    skus = set()
    for p in products:
        s = p["SKU"]
        if s in skus:
            issues.append(f"DUP SKU: {s}")
        skus.add(s)
        if p.get("Categories") == "Accessories" and "mattress" in p.get("Name", "").lower():
            issues.append(f"MATTRESS IN ACCESSORIES: {s}")
        if p.get("Published") == 1 and not p.get("Regular price"):
            issues.append(f"NO PRICE (published): {s}")
        if not p.get("Images"):
            issues.append(f"NO IMAGE: {s}")
    return issues


# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    global IMAGE_BASE_URL

    args = parse_args()
    pdf_path = args.pdf.expanduser()
    images_root = args.images_root.expanduser()
    output_dir = args.output_dir.expanduser()
    output_dir.mkdir(parents=True, exist_ok=True)
    IMAGE_BASE_URL = args.image_base_url.rstrip("/")

    if not pdf_path.is_file():
        raise FileNotFoundError(f"PDF file not found: {pdf_path}")
    if not images_root.is_dir():
        log.warning("Image folder not found: %s", images_root)

    log.info("=" * 60)
    log.info("Pierre Cardin Bedding — WooCommerce CSV Generator v2")
    log.info("=" * 60)

    log.info("[1/5] Parsing PDF...")
    pdf_models = parse_pdf(pdf_path)
    log.info(f"  {len(pdf_models)} model blocks")
    for m in pdf_models:
        log.info(f"    {m['name']:30s}  col={m['collection'] or '?':10s}  sizes={len(m['sizes'])}  acc={list(m['accessories'].keys())}")

    log.info("[2/5] Scanning images...")
    image_map, ai_bgs, yatak = scan_images(images_root)
    log.info(f"  {len(image_map)} folders, {len(ai_bgs)} AI bgs, {len(yatak)} yatak labels")

    log.info("[3/5] Building products...")
    products, mi, mp, mr, im = build_all(pdf_models, image_map, ai_bgs, yatak)

    log.info("[4/5] Validating...")
    issues = validate(products)
    for iss in issues[:30]:
        log.warning(f"  {iss}")
    if len(issues) > 30:
        log.warning(f"  ...+{len(issues)-30} more")

    log.info("[5/5] Exporting...")
    export_csv(products, output_dir / "final_import.csv")
    export_csv(products, output_dir / "final_import_semicolon.csv", delim=";")
    export_xlsx(products, mi, mp, mr, im, output_dir / "final_import.xlsx")
    write_report_csv(im, output_dir / "image_mapping.csv")
    write_report_csv(mr, output_dir / "manual_review.csv")
    write_report_csv(mp, output_dir / "missing_prices.csv")
    write_report_csv(mi, output_dir / "missing_images.csv")

    # ── log.txt ─────────────────────────────────────────────────────
    cats = defaultdict(int)
    for p in products:
        cats[p.get("Categories", "?")] += 1
    no_img = sum(1 for p in products if not p.get("Images"))
    no_prc = sum(1 for p in products if p.get("Published") == 1 and not p.get("Regular price"))

    with open(output_dir / "log.txt", "w", encoding="utf-8") as f:
        f.write("Pierre Cardin Bedding — Generation Log\n" + "=" * 50 + "\n\n")
        f.write(f"Total products: {len(products)}\n")
        for c, n in sorted(cats.items()):
            f.write(f"  {c}: {n}\n")
        f.write(f"\nWithout images: {no_img}\nWithout price (published): {no_prc}\n")
        f.write(f"Manual review: {len(mr)}\nMissing prices: {len(mp)}\n\n")
        f.write("Validation issues:\n")
        for iss in issues:
            f.write(f"  - {iss}\n")

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Total products:   {len(products)}")
    for c in ["Beds", "Bedroom Sets", "Mattresses", "Accessories"]:
        print(f"  {c:17s} {cats.get(c, 0)}")
    print(f"Without images:   {no_img}")
    print(f"Without price:    {no_prc}")
    print(f"Manual review:    {len(mr)}")
    print(f"Validation issues: {len(issues)}")
    print(f"\nOutput -> {output_dir}")
    print("=" * 60)


if __name__ == "__main__":
    main()
